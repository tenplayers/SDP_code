from GKP_ampDamp import *
import openpyxl
import datetime
file_name = str(datetime.datetime.now())+'.xlsx'

# to open the excel sheet and if it has macros
srcfile = openpyxl.load_workbook(filename=file_name, read_only=False, keep_vba=True)

# get sheetname from the file
sheetname = srcfile['Sheet1']

row = 2
for gamma in np.linspace(0.01,0.1,10):
    for Delta in np.linspace(0.221, 0.481,6):
        print('--- gamma =',gamma,'--- Delta =',Delta)
        dimL=20
        cutoff = 5
        ElmuBasis = GKP_ElmuBasis(Delta = Delta, gamma = gamma, m_sum_cutoff=20,M_sum_cutoff=5,l_cut=20)
        n_Delta = ElmuBasis.n_Delta
        # check
        if 1:
            ck = check_basis(ElmuBasis = ElmuBasis,nBasis = None)
            ck.trM()

        # do optimization
        infid_M = ElmuBasis.tranpose_fid()
        print('Elmu:',infid_M)
        # write to row 1,col 1 explicitly, this type of writing is useful to
        # write something in loops
        sheetname.cell(row=row, column=1).value = Delta
        sheetname.cell(row=row, column=2).value = n_Delta
        sheetname.cell(row=row, column=3).value = gamma
        sheetname.cell(row=row, column=4).value = infid_M

        # save it as a new file, the original file is untouched and here I am saving
        # it as xlsm(m here denotes macros).
        srcfile.save(file_name)
        row += 1
